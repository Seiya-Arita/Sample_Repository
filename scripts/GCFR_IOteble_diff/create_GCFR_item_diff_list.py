# -----------------------------------------------------------------------------
# 処理概要:
# このスクリプトは、指定フォルダ配下にある全てのマッピング定義書Excelファイルの
# 「マッピング定義書」シートを走査し、入力項目・出力項目の型変換（特に文字型→数値型）を検出、
# その対応一覧を集計してExcelファイルに出力するツールです。
#
# 主な処理内容:
#   1. 指定フォルダ内の全Excelファイルを検索し、"マッピング定義書"シートから
#      ジョブ名・入出力テーブル情報・項目マッピング情報を抽出
#   2. 入力・出力テーブルのカラム情報（型や桁数など）をSQL Serverから取得
#   3. 入力カラムがvarchar/char型、出力カラムがdecimal/int型となる型変換項目を抽出
#   4. ジョブごとに変換項目の詳細を新規シートに出力し、全体サマリシートも作成
#   5. 結果を1つのExcelファイル（result.xlsx）として保存
#
# 注意事項:
#   - SQL Serverへの接続情報は config.py の DB_CONN_STR で指定します
#   - Excelレイアウト（シート名・セル位置）が変更された場合は該当箇所の修正が必要です
#   - 同名ジョブが複数存在する場合、出力Excelのシート名重複に注意してください
#   - 項目名や型の取得はINFORMATION_SCHEMA.COLUMNSを参照するため、事前にDBにテーブル定義があること
# -----------------------------------------------------------------------------

import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pyodbc
from openpyxl import load_workbook
from config import DB_CONN_STR, FOLDER_PATH, INPUT_COLUMNS_START_CELL, INPUT_TABLE_CELL, JOB_NAME_CELL, OUTPUT_EXCEL, OUTPUT_SCHEMA_CELL, OUTPUT_TABLE_CELL, SHEET_NAME_TARGET

conn = pyodbc.connect(DB_CONN_STR)

wb_result = Workbook()
ws_summary = wb_result.active
ws_summary.title = "ジョブ一覧"
ws_summary.append(["ジョブ名", "入力", "出力"])

for root, dirs, files in os.walk(FOLDER_PATH):
    for file in files:
        if file.endswith('.xlsx') and not file.startswith('~$'):
            file_path = os.path.join(root, file)
            try:
                wb = load_workbook(file_path, data_only=True)
                ws = wb[SHEET_NAME_TARGET]

                job_name = ws[JOB_NAME_CELL].value
                input_schema = 'HN_INFO'
                input_table = ws[INPUT_TABLE_CELL].value
                output_schema = ws[OUTPUT_SCHEMA_CELL].value
                if output_schema == '${DB_T_SEM}':
                    output_schema = 'HN_T_SEM'
                else:
                    output_schema = 'HN_T_SRC'
                output_table = ws[OUTPUT_TABLE_CELL].value

                # 項目取得
                columns = []
                row_num = int(INPUT_COLUMNS_START_CELL[1:])
                while True:
                    input_col = ws[f"B{row_num}"].value
                    output_col = ws[f"G{row_num}"].value
                    if not input_col or not output_col:
                        break
                    columns.append((input_col.strip(), output_col.strip()))
                    row_num += 1

                result_rows = []
                for input_col, output_col in columns:
                    query_in = f"""
                        SELECT 
                            a.COLUMN_NAME
                            , a.DATA_TYPE
                            , a.CHARACTER_MAXIMUM_LENGTH
                            , a.NUMERIC_PRECISION
                            , a.NUMERIC_SCALE
                        from
                            INFORMATION_SCHEMA.COLUMNS a
                        where
                            a.COLUMN_NAME = '{input_col}'
                            and a.TABLE_NAME = '{input_table}'
                            and a.TABLE_SCHEMA = '{input_schema}'
                    """
                    query_out = f"""
                        SELECT 
                            a.COLUMN_NAME
                            , a.DATA_TYPE
                            , a.CHARACTER_MAXIMUM_LENGTH
                            , a.NUMERIC_PRECISION
                            , a.NUMERIC_SCALE
                        from
                            INFORMATION_SCHEMA.COLUMNS a
                        where
                            a.COLUMN_NAME = '{output_col}'
                            and a.TABLE_NAME = '{output_table}'
                            and a.TABLE_SCHEMA = '{output_schema}'
                    """

                    df_in = pd.read_sql(query_in, conn)
                    df_out = pd.read_sql(query_out, conn)

                    if not df_in.empty and not df_out.empty:
                        if (df_in['DATA_TYPE'][0] == 'varchar' or df_in['DATA_TYPE'][0] == 'char') and (df_out['DATA_TYPE'][0] == 'decimal' or df_out['DATA_TYPE'][0] == 'int'):
                            row_in = df_in.iloc[0].to_list()
                            row_out = df_out.iloc[0].to_list()
                            result_rows.append(row_in + row_out)

                if result_rows:
                    columns = [f"input_{col}" for col in df_in.columns] + [f"output_{col}" for col in df_out.columns]
                    merged_df = pd.DataFrame(result_rows, columns=columns)

                    ws_job = wb_result.create_sheet(job_name)
                    for r in dataframe_to_rows(merged_df, index=False, header=True):
                        ws_job.append(r)

                    ws_summary.append([job_name, input_table, output_table])
                print(job_name)

            except Exception as e:
                print(f"[エラー] {file}: {e}")
    break

wb_result.save(OUTPUT_EXCEL)
print(f"[完了] {OUTPUT_EXCEL} に出力しました。")
