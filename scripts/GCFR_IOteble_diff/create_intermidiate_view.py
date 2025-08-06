# -----------------------------------------------------------------------------
# 処理概要:
# 本スクリプトは、SQL Serverのテーブル定義および拡張プロパティ（説明文）を参照し、
# 指定されたExcelファイルの内容（ジョブ一覧・各ジョブシート）をもとに、
# 中間View作成用のSQLCMDスクリプト（.cvファイル）を自動生成します。
#
# 主な処理内容:
#  1. Excelファイル（ジョブ一覧・各ジョブシート）から、対象のテーブル名・View名・CASE対象カラム名を取得
#  2. データベースから各テーブル・カラムの説明文（拡張プロパティ）を取得
#  3. テーブル定義・CASE条件・説明情報をもとに、View作成および説明付与のSQLCMDスクリプトを生成
#  4. 各Viewごとにスクリプトファイル（.cv）として出力
#
# 対象ファイル:
#   - 入力:  Excelファイル（EXCEL_PATHにて定義）
#   - 出力:  View作成用SQLCMDスクリプト（View_DDLs配下の各 .cv ファイル）
#
# 注意事項:
#   - SQL Serverへの接続情報は config.py の DB_CONN_STR で指定します
#   - ExcelのフォーマットやDBスキーマ構成が変わった場合は、該当箇所の修正が必要です
#   - View名・テーブル名・カラム名に日本語や特殊文字が含まれる場合、SQL生成・実行時に注意してください
# -----------------------------------------------------------------------------

import pyodbc
import openpyxl
import os
from config import DB_CONN_STR, EXCEL_PATH, OUTPUT_DIR

def get_columns_with_descriptions(table_name, schema='HN_INFO'):
    with pyodbc.connect(DB_CONN_STR) as conn:
        cursor = conn.cursor()
        
        query = f"""
        SELECT 
            c.COLUMN_NAME,
            ep.value AS COLUMN_DESCRIPTION
        FROM INFORMATION_SCHEMA.COLUMNS c
        LEFT JOIN sys.extended_properties ep
            ON ep.major_id = OBJECT_ID(QUOTENAME('{schema}') + '.' + QUOTENAME('{table_name}'))
            AND ep.minor_id = 
                (SELECT column_id FROM sys.columns 
                 WHERE object_id = OBJECT_ID(QUOTENAME('{schema}') + '.' + QUOTENAME('{table_name}'))
                 AND name = c.COLUMN_NAME)
            AND ep.name = 'MS_Description'
        WHERE c.TABLE_SCHEMA = ? AND c.TABLE_NAME = ?
        ORDER BY c.ORDINAL_POSITION
        """
        
        cursor.execute(query, (schema, table_name))
        rows = cursor.fetchall()
        return [(row.COLUMN_NAME, row.COLUMN_DESCRIPTION or '') for row in rows]

def get_table_description(table_name, schema='HN_INFO'):
    with pyodbc.connect(DB_CONN_STR) as conn:
        cursor = conn.cursor()

        query = f"""
        SELECT ep.value AS TABLE_DESCRIPTION
        FROM sys.extended_properties ep
        WHERE ep.major_id = OBJECT_ID(QUOTENAME('{schema}') + '.' + QUOTENAME('{table_name}'))
        AND ep.minor_id = 0
        AND ep.name = 'MS_Description'
        """
        cursor.execute(query)
        row = cursor.fetchone()
        return row.TABLE_DESCRIPTION if row else ""
    
def generate_sqlcmd_script(columns_with_desc, case_columns, view_name, table_name):
    item_list = ',\n    '.join([f"[{col}]" for col, _ in columns_with_desc])
    select_item_list = []
    for col, desc in columns_with_desc:
        if col in case_columns:
            select_item_list.append(f"CASE \n        WHEN TRIM([{col}]) = '' THEN '0' \n        ELSE [{col}] \n    END AS [{col}]")
        else:
            select_item_list.append(f'[{col}]')
    select_list = ',\n    '.join(select_item_list)

    sql_lines = [
        'sqlcmd -S ${SQL_SERVER} -d ${SQL_DB} -U ${SQL_USER} -P ${SQL_PASS} -f ${CHARSET} -t 0 -e -b -N o -Q "',
        "DROP VIEW IF EXISTS ${INFO_DB}." + view_name + " ;",
        "GO",
        "CREATE VIEW ${INFO_DB}." + view_name,
        "(",
        "    " + item_list,
        ")",
        "AS",
        "SELECT",
        "    " + select_list,
        "FROM ${INFO_DB}." + table_name,
        "GO"
    ]

    table_description = get_table_description(table_name)
    description_clean = table_description.replace("'", "''")
    sql_lines.append(
        "EXEC sys.sp_addextendedproperty @name = N'MS_Description', "
        f"@value = N'{description_clean}', "
        "@level0type = N'SCHEMA', @level0name = '${INFO_DB}', "
        "@level1type = N'VIEW', @level1name = '" + view_name + "';"
    )

    for col, desc in columns_with_desc:
        description_clean = desc.replace("'", "''")
        sql_lines.append(
            "EXEC sys.sp_addextendedproperty @name = N'MS_Description', "
            f"@value = N'{description_clean}', "
            "@level0type = N'SCHEMA', @level0name = '${INFO_DB}', "
            "@level1type = N'VIEW', @level1name = '" + view_name + "', "
            "@level2type = N'COLUMN', @level2name = '" + col + "';"
        )

    sql_lines.append('"')
    sql_lines.extend([
        "if [ $? -ne 0 ]; then",
        "  exit 8",
        "else",
        "  exit 0",
        "fi"
    ])

    return "\n".join(sql_lines)

def get_target_tables_and_jobs(excel_path):
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb['ジョブ一覧']

    result = []
    for row in ws.iter_rows(min_row=2, values_only=True):  # ヘッダー除外
        job_name, table_name, view_name = row[0], row[1], row[3]
        if table_name and table_name.startswith("T_"):
            result.append((job_name, table_name, view_name))
    return result

def get_item_list_from_job_sheet(wb, job_name):
    if job_name not in wb.sheetnames:
        return []

    ws = wb[job_name]
    items = []
    for row in ws.iter_rows(min_row=2, values_only=True):  # ヘッダー除外
        item = row[0]
        if item:
            items.append(item)
    return items

def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    target_list = get_target_tables_and_jobs(EXCEL_PATH)
    for job_name, table_name, view_name in target_list:
        columns_with_desc = get_columns_with_descriptions(table_name)
        case_columns = get_item_list_from_job_sheet(wb, job_name)

        script = generate_sqlcmd_script(columns_with_desc, case_columns, view_name, table_name)

        with open(f'{OUTPUT_DIR}/{view_name}.cv', 'w', encoding='utf-8', newline='\n') as f:
            f.write(script)

        print(f"{view_name}.cv を出力しました。")

if __name__ == "__main__":
    main()
